from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from threading import Lock

from openpyxl import Workbook, load_workbook


@dataclass(frozen=True)
class ExcelExporter:
    export_dir: Path
    prefix: str = "zproveart"

    def __post_init__(self):
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def daily_path(self, day: date | None = None) -> Path:
        day = day or date.today()
        return self.export_dir / f"{self.prefix}_{day.strftime('%Y%m%d')}.xlsx"


_lock = Lock()


def append_row_daily(exporter: ExcelExporter, itmref: str, selected: bool, comment: str) -> Path:
    filepath = exporter.daily_path()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row = [ts, itmref, "1" if selected else "0", comment]

    with _lock:
        if filepath.exists():
            wb = load_workbook(filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "ZPROVEART"
            ws.append(["timestamp", "itmref", "selected", "comment"])

        ws.append(row)
        wb.save(filepath)

    return filepath
